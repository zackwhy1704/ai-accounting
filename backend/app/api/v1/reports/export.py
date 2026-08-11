"""
Generic XLSX export for reports. The frontend already holds every rendered
report as headers + rows (the same data its CSV export uses); this turns that
into a styled .xlsx download so any report gains Excel export with one call.
"""
import io

from fastapi import APIRouter, Depends, Response
from pydantic import BaseModel, Field

from app.core.security import get_current_user

router = APIRouter()


class XlsxExportRequest(BaseModel):
    filename: str = "report"
    sheet_name: str = "Report"
    title: str | None = None
    headers: list[str]
    rows: list[list] = Field(default_factory=list)


def build_xlsx(payload: XlsxExportRequest) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (payload.sheet_name or "Report")[:31]

    row_idx = 1
    if payload.title:
        ws.cell(row=1, column=1, value=payload.title).font = Font(bold=True, size=13)
        row_idx = 3

    for col, header in enumerate(payload.headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.font = Font(bold=True)

    for r, row in enumerate(payload.rows, start=row_idx + 1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    for col in range(1, len(payload.headers) + 1):
        width = max(
            [len(str(payload.headers[col - 1]))]
            + [len(str(row[col - 1])) for row in payload.rows if len(row) >= col]
        ) if payload.rows or payload.headers else 10
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/export-xlsx")
async def export_xlsx(payload: XlsxExportRequest, current_user: dict = Depends(get_current_user)):
    content = build_xlsx(payload)
    safe_name = "".join(ch for ch in payload.filename if ch.isalnum() or ch in ("-", "_", " ")).strip() or "report"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
