"""XLSX export builder — produces a valid workbook with headers/rows/title."""
import io

from openpyxl import load_workbook

from app.api.v1.reports.export import XlsxExportRequest, build_xlsx


def test_build_xlsx_roundtrip():
    payload = XlsxExportRequest(
        filename="pl", sheet_name="P&L Report", title="Profit & Loss — Aug 2026",
        headers=["Account", "Amount"],
        rows=[["Sales Revenue", 1000.5], ["Rent", -200]],
    )
    wb = load_workbook(io.BytesIO(build_xlsx(payload)))
    ws = wb.active
    assert ws.title == "P&L Report"
    assert ws.cell(row=1, column=1).value == "Profit & Loss — Aug 2026"
    assert ws.cell(row=3, column=1).value == "Account"
    assert ws.cell(row=4, column=2).value == 1000.5
    assert ws.cell(row=5, column=1).value == "Rent"


def test_build_xlsx_no_title_headers_first_row():
    payload = XlsxExportRequest(headers=["A"], rows=[[1]])
    wb = load_workbook(io.BytesIO(build_xlsx(payload)))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=2, column=1).value == 1
